"""
Exportação PREMIUM - Excel e PDF com Design Profissional
"""
from datetime import datetime
from decimal import Decimal
import io
from gerenciador_logos import obter_logo_como_bytes

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
except:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
except:
    import subprocess
    subprocess.check_call(['pip', 'install', 'reportlab'])
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT


def gerar_excel_ficha(ficha, output_stream=None):
    """Gera Excel PREMIUM com design profissional"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Técnica"
    
    # CORES PREMIUM
    COR_PRINCIPAL = "667EEA"
    COR_SECUNDARIA = "764BA2"
    COR_DESTAQUE = "10B981"
    
    # Larguras
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    linha = 1
    
    # CABEÇALHO COM FUNDO COLORIDO
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_PRINCIPAL, end_color=COR_PRINCIPAL, fill_type="solid")
        ws[f'{col}{linha+1}'].fill = PatternFill(start_color=COR_PRINCIPAL, end_color=COR_PRINCIPAL, fill_type="solid")
    
    ws.row_dimensions[linha].height = 25
    ws.row_dimensions[linha+1].height = 25
    
    # Logo
    if ficha.cliente.logo_path:
        try:
            logo_bytes = obter_logo_como_bytes(ficha.cliente.logo_path, 'PNG')
            if logo_bytes:
                img = XLImage(io.BytesIO(logo_bytes))
                img.width = 100
                img.height = 60
                ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Erro logo: {e}")
    
    # Título
    ws.merge_cells('C1:E2')
    titulo = ws['C1']
    titulo.value = "FICHA TÉCNICA"
    titulo.font = Font(bold=True, size=18, color="FFFFFF")
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    linha = 4
    
    # INFO - Card cinza claro
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{linha}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws[f'{col}{linha+1}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws[f'{col}{linha+2}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    ws['B4'] = "Cliente:"
    ws['B4'].font = Font(bold=True, size=11)
    ws.merge_cells('C4:E4')
    ws['C4'] = ficha.cliente.nome
    ws['C4'].font = Font(size=11)
    
    ws['B5'] = "Código:"
    ws['B5'].font = Font(bold=True)
    ws['C5'] = ficha.codigo
    ws['D5'] = "Data:"
    ws['D5'].font = Font(bold=True)
    ws['E5'] = ficha.data_criacao.strftime('%d/%m/%Y')
    
    ws['B6'] = "Receita:"
    ws['B6'].font = Font(bold=True)
    ws.merge_cells('C6:E6')
    ws['C6'] = ficha.nome
    ws['C6'].font = Font(bold=True, size=12)
    
    linha = 8
    
    # CABEÇALHO TABELA
    headers = [('B', 'INGREDIENTE'), ('C', 'QUANTIDADE'), ('D', 'PREÇO UNIT.'), ('E', 'SUBTOTAL')]
    for col, text in headers:
        cell = ws[f'{col}{linha}']
        cell.value = text
        cell.fill = PatternFill(start_color=COR_SECUNDARIA, end_color=COR_SECUNDARIA, fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[linha].height = 25
    
    # INGREDIENTES
    for idx, item in enumerate(ficha.itens):
        linha += 1
        cor_linha = "FFFFFF" if idx % 2 == 0 else "F9FAFB"
        
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{linha}'].fill = PatternFill(start_color=cor_linha, end_color=cor_linha, fill_type="solid")
        
        # Suporte a fichas aninhadas
        if item.tipo_item == 'ficha' and item.ficha_ingrediente:
            # Item é uma ficha técnica
            ws[f'B{linha}'] = f"📋 {item.ficha_ingrediente.nome}"
            ws[f'C{linha}'] = f"{float(item.quantidade)} g"
            ws[f'C{linha}'].alignment = Alignment(horizontal='center')
            custo_por_grama = float(item.custo_unitario_historico) if item.custo_unitario_historico else 0
            ws[f'D{linha}'] = custo_por_grama
            ws[f'D{linha}'].number_format = 'R$ #,##0.0000'
            ws[f'D{linha}'].alignment = Alignment(horizontal='right')
            ws[f'E{linha}'] = float(item.custo_item)
            ws[f'E{linha}'].number_format = 'R$ #,##0.00'
            ws[f'E{linha}'].alignment = Alignment(horizontal='right')
            ws[f'E{linha}'].font = Font(bold=True)
        elif item.insumo:
            # Item é um insumo normal
            ws[f'B{linha}'] = item.insumo.nome
            ws[f'C{linha}'] = f"{float(item.quantidade)} {item.insumo.unidade_medida}"
            ws[f'C{linha}'].alignment = Alignment(horizontal='center')
            ws[f'D{linha}'] = float(item.insumo.preco_unitario)
            ws[f'D{linha}'].number_format = 'R$ #,##0.00'
            ws[f'D{linha}'].alignment = Alignment(horizontal='right')
            ws[f'E{linha}'] = float(item.custo_item)
            ws[f'E{linha}'].number_format = 'R$ #,##0.00'
            ws[f'E{linha}'].alignment = Alignment(horizontal='right')
            ws[f'E{linha}'].font = Font(bold=True)
    
    # TOTAL
    linha += 2
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_DESTAQUE, end_color=COR_DESTAQUE, fill_type="solid")
    
    ws.merge_cells(f'B{linha}:D{linha}')
    ws[f'B{linha}'] = "CUSTO TOTAL DOS INSUMOS"
    ws[f'B{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'B{linha}'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws[f'E{linha}'] = float(ficha.custo_insumos)
    ws[f'E{linha}'].number_format = 'R$ #,##0.00'
    ws[f'E{linha}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'E{linha}'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[linha].height = 30
    
    # PREÇO VENDA
    if ficha.preco_venda > 0:
        linha += 1
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_PRINCIPAL, end_color=COR_PRINCIPAL, fill_type="solid")
        
        ws.merge_cells(f'B{linha}:D{linha}')
        ws[f'B{linha}'] = f"PREÇO DE VENDA (Margem: {float(ficha.margem_percentual):.0f}%)"
        ws[f'B{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'B{linha}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'E{linha}'] = float(ficha.preco_venda)
        ws[f'E{linha}'].number_format = 'R$ #,##0.00'
        ws[f'E{linha}'].font = Font(bold=True, size=16, color="FFFFFF")
        ws[f'E{linha}'].alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[linha].height = 35
    
    # RODAPÉ
    linha += 3
    ws.merge_cells(f'B{linha}:E{linha}')
    ws[f'B{linha}'] = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} | Sistema Ficha Técnica PRO"
    ws[f'B{linha}'].font = Font(size=9, color="6B7280", italic=True)
    ws[f'B{linha}'].alignment = Alignment(horizontal='center')
    
    # Salvar
    if output_stream is None:
        output_stream = io.BytesIO()
    
    try:
        wb.save(output_stream)
        if isinstance(output_stream, io.BytesIO):
            output_stream.seek(0)
            return output_stream.getvalue()
    except Exception as e:
        print(f"Erro ao salvar Excel: {e}")
        return None
    
    return None


def gerar_pdf_ficha(ficha, output_stream=None):
    """Gera PDF PREMIUM com design profissional"""
    
    if output_stream is None:
        output_stream = io.BytesIO()
    
    doc = SimpleDocTemplate(output_stream, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    elementos = []
    
    # CABEÇALHO com logo
    if ficha.cliente.logo_path:
        try:
            logo_bytes = obter_logo_como_bytes(ficha.cliente.logo_path, 'PNG')
            if logo_bytes:
                img = RLImage(io.BytesIO(logo_bytes), width=5*cm, height=3*cm, kind='proportional')
                titulo = Paragraph("<b>FICHA TÉCNICA</b>", ParagraphStyle('Title', fontSize=24, textColor=colors.HexColor('#667eea'), alignment=TA_CENTER))
                header = Table([[img, titulo]], colWidths=[6*cm, 10*cm])
                header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (1,0), (1,0), 'CENTER')]))
                elementos.append(header)
        except:
            elementos.append(Paragraph("<b>FICHA TÉCNICA</b>", ParagraphStyle('T', fontSize=24, textColor=colors.HexColor('#667eea'), alignment=TA_CENTER)))
    else:
        elementos.append(Paragraph("<b>FICHA TÉCNICA</b>", ParagraphStyle('T', fontSize=24, textColor=colors.HexColor('#667eea'), alignment=TA_CENTER)))
    
    elementos.append(Spacer(1, 0.8*cm))
    
    # INFO
    info_data = [
        [Paragraph("<b>Cliente:</b>", styles['Normal']), Paragraph(ficha.cliente.nome, styles['Normal']), '', ''],
        [Paragraph("<b>Código:</b>", styles['Normal']), ficha.codigo, Paragraph("<b>Data:</b>", styles['Normal']), ficha.data_criacao.strftime('%d/%m/%Y')],
        [Paragraph("<b>Receita:</b>", styles['Normal']), Paragraph(f"<b>{ficha.nome}</b>", ParagraphStyle('B', fontSize=14, textColor=colors.HexColor('#667eea'))), '', '']
    ]
    
    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 2.5*cm, 4.5*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F9FAFB')),
        ('SPAN', (1,0), (3,0)),
        ('SPAN', (1,2), (3,2)),
        ('PADDING', (0,0), (-1,-1), 12),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elementos.append(info_table)
    elementos.append(Spacer(1, 1*cm))
    
    # INGREDIENTES
    ing_data = [[Paragraph("<b>INGREDIENTE</b>", styles['Normal']), Paragraph("<b>QTD</b>", styles['Normal']), 
                 Paragraph("<b>PREÇO</b>", styles['Normal']), Paragraph("<b>SUBTOTAL</b>", styles['Normal'])]]
    
    for item in ficha.itens:
        # Suporte a fichas aninhadas
        if item.tipo_item == 'ficha' and item.ficha_ingrediente:
            # Item é uma ficha técnica
            nome_display = f"📋 {item.ficha_ingrediente.nome}"
            qtd_display = f"{float(item.quantidade)} g"
            preco_display = f"R$ {float(item.custo_unitario_historico):,.4f}/g" if item.custo_unitario_historico else "R$ 0,00"
        elif item.insumo:
            # Item é um insumo normal
            nome_display = item.insumo.nome
            qtd_display = f"{float(item.quantidade)} {item.insumo.unidade_medida}"
            preco_display = f"R$ {float(item.insumo.preco_unitario):,.2f}"
        else:
            continue
        
        ing_data.append([
            nome_display,
            qtd_display,
            preco_display,
            Paragraph(f"<b>R$ {float(item.custo_item):,.2f}</b>", styles['Normal'])
        ])
    
    ing_table = Table(ing_data, colWidths=[7*cm, 3*cm, 3*cm, 3*cm])
    ing_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#764BA2')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('ALIGN', (3,1), (3,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('PADDING', (0,0), (-1,-1), 10),
    ]))
    elementos.append(ing_table)
    elementos.append(Spacer(1, 0.8*cm))
    
    # TOTAIS
    totais_data = [[Paragraph("<b>CUSTO TOTAL</b>", ParagraphStyle('B', fontSize=14, textColor=colors.white)), 
                    Paragraph(f"<b>R$ {float(ficha.custo_insumos):,.2f}</b>", ParagraphStyle('B', fontSize=16, textColor=colors.white, alignment=TA_RIGHT))]]
    
    if ficha.preco_venda > 0:
        totais_data.append([Paragraph(f"<b>PREÇO DE VENDA</b>", ParagraphStyle('B', fontSize=14, textColor=colors.white)),
                           Paragraph(f"<b>R$ {float(ficha.preco_venda):,.2f}</b>", ParagraphStyle('B', fontSize=18, textColor=colors.white, alignment=TA_RIGHT))])
    
    totais_table = Table(totais_data, colWidths=[10*cm, 6*cm])
    totais_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#10B981')),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#667EEA')),
        ('PADDING', (0,0), (-1,-1), 15),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elementos.append(totais_table)
    
    # RODAPÉ
    elementos.append(Spacer(1, 2*cm))
    elementos.append(Paragraph(f"<font size=9 color='#6B7280'>Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} | Sistema Ficha Técnica PRO</font>", 
                               ParagraphStyle('F', alignment=TA_CENTER)))
    
    doc.build(elementos)
    if isinstance(output_stream, io.BytesIO):
        output_stream.seek(0)
        return output_stream.getvalue()
    return None
