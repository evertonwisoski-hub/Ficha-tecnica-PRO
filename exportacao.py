"""
Módulo de Exportação de Fichas Técnicas
Gera arquivos Excel (XLS) e PDF
"""
from datetime import datetime
from decimal import Decimal
import io
from gerenciador_logos import obter_logo_como_bytes

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'reportlab'])
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


def gerar_excel_ficha(ficha, output_stream=None):
    """Gera arquivo Excel de uma ficha técnica"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Técnica"
    
    # Estilos
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, size=14, color="FFFFFF")
    title_font = Font(bold=True, size=16)
    bold_font = Font(bold=True)
    
    # Larguras
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    linha = 1
    
    # Logo se existir
    if ficha.cliente.logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_bytes = obter_logo_como_bytes(ficha.cliente.logo_path, 'PNG')
            if logo_bytes:
                logo_stream = io.BytesIO(logo_bytes)
                img = XLImage(logo_stream)
                img.width = 100
                img.height = 60
                ws.add_image(img, 'B1')
                ws.row_dimensions[1].height = 60
        except:
            pass
    
    # Título
    ws.merge_cells(f'C{linha}:E{linha}')
    ws[f'C{linha}'].value = "FICHA TÉCNICA"
    ws[f'C{linha}'].font = title_font
    ws[f'C{linha}'].alignment = Alignment(horizontal='center')
    
    linha = 3
    ws[f'B{linha}'] = "Código:"
    ws[f'B{linha}'].font = bold_font
    ws[f'C{linha}'] = ficha.codigo
    ws[f'D{linha}'] = "Data:"
    ws[f'D{linha}'].font = bold_font
    ws[f'E{linha}'] = ficha.data_criacao.strftime('%d/%m/%Y')
    
    linha += 1
    ws[f'B{linha}'] = "Receita:"
    ws[f'B{linha}'].font = bold_font
    ws.merge_cells(f'C{linha}:E{linha}')
    ws[f'C{linha}'] = ficha.nome
    
    linha += 1
    ws[f'B{linha}'] = "Cliente:"
    ws[f'B{linha}'].font = bold_font
    ws.merge_cells(f'C{linha}:E{linha}')
    ws[f'C{linha}'] = ficha.cliente.nome
    
    linha += 2
    # Cabeçalho ingredientes
    for col, text in [('B', 'INGREDIENTE'), ('C', 'QUANTIDADE'), ('D', 'PREÇO'), ('E', 'CUSTO')]:
        ws[f'{col}{linha}'].value = text
        ws[f'{col}{linha}'].fill = header_fill
        ws[f'{col}{linha}'].font = header_font
        ws[f'{col}{linha}'].alignment = Alignment(horizontal='center')
    
    # Ingredientes
    for item in ficha.itens:
        linha += 1
        ws[f'B{linha}'] = item.insumo.nome
        ws[f'C{linha}'] = f"{float(item.quantidade)} {item.insumo.unidade_medida}"
        ws[f'D{linha}'] = f"R$ {float(item.insumo.preco_unitario):,.2f}"
        ws[f'E{linha}'] = f"R$ {float(item.custo_item):,.2f}"
    
    # Total
    linha += 2
    ws[f'B{linha}'] = "CUSTO TOTAL"
    ws[f'B{linha}'].font = bold_font
    ws[f'E{linha}'] = f"R$ {float(ficha.custo_insumos):,.2f}"
    ws[f'E{linha}'].font = bold_font
    ws[f'E{linha}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    if ficha.preco_venda > 0:
        linha += 1
        ws[f'B{linha}'] = f"PREÇO DE VENDA"
        ws[f'B{linha}'].font = Font(bold=True, color="1B5E20")
        ws[f'E{linha}'] = f"R$ {float(ficha.preco_venda):,.2f}"
        ws[f'E{linha}'].font = Font(bold=True, size=12, color="1B5E20")
        ws[f'E{linha}'].fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    
    # Salvar
    if output_stream is None:
        output_stream = io.BytesIO()
    
    wb.save(output_stream)
    
    if isinstance(output_stream, io.BytesIO):
        output_stream.seek(0)
        return output_stream.getvalue()
    
    return None


def gerar_pdf_ficha(ficha, output_stream=None):
    """Gera arquivo PDF de uma ficha técnica"""
    
    if output_stream is None:
        output_stream = io.BytesIO()
    
    doc = SimpleDocTemplate(output_stream, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#4CAF50'), 
                                  spaceAfter=30, alignment=TA_CENTER, fontName='Helvetica-Bold')
    
    elementos = []
    
    # Logo e título
    if ficha.cliente.logo_path:
        try:
            from reportlab.platypus import Image as RLImage
            logo_bytes = obter_logo_como_bytes(ficha.cliente.logo_path, 'PNG')
            if logo_bytes:
                logo_stream = io.BytesIO(logo_bytes)
                img = RLImage(logo_stream, width=4*cm, height=2.5*cm, kind='proportional')
                header_table = Table([[img, Paragraph("FICHA TÉCNICA", title_style)]], colWidths=[5*cm, 11*cm])
                header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('ALIGN', (1, 0), (1, 0), 'CENTER')]))
                elementos.append(header_table)
                elementos.append(Spacer(1, 0.5*cm))
            else:
                elementos.append(Paragraph("FICHA TÉCNICA", title_style))
                elementos.append(Spacer(1, 0.5*cm))
        except:
            elementos.append(Paragraph("FICHA TÉCNICA", title_style))
            elementos.append(Spacer(1, 0.5*cm))
    else:
        elementos.append(Paragraph("FICHA TÉCNICA", title_style))
        elementos.append(Spacer(1, 0.5*cm))
    
    # Informações
    info_data = [
        ['Código:', ficha.codigo, 'Data:', ficha.data_criacao.strftime('%d/%m/%Y')],
        ['Receita:', ficha.nome, '', ''],
        ['Cliente:', ficha.cliente.nome, '', '']
    ]
    
    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('SPAN', (1, 1), (3, 1)),
        ('SPAN', (1, 2), (3, 2)),
    ]))
    
    elementos.append(info_table)
    elementos.append(Spacer(1, 1*cm))
    
    # Ingredientes
    ingredientes_data = [['INGREDIENTE', 'QUANTIDADE', 'PREÇO', 'CUSTO']]
    for item in ficha.itens:
        ingredientes_data.append([
            item.insumo.nome,
            f"{float(item.quantidade)} {item.insumo.unidade_medida}",
            f"R$ {float(item.insumo.preco_unitario):,.2f}",
            f"R$ {float(item.custo_item):,.2f}"
        ])
    
    ingredientes_table = Table(ingredientes_data, colWidths=[7*cm, 3*cm, 3*cm, 3*cm])
    ingredientes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    elementos.append(ingredientes_table)
    elementos.append(Spacer(1, 0.8*cm))
    
    # Resumo
    resumo_data = [['CUSTO TOTAL', f"R$ {float(ficha.custo_insumos):,.2f}"]]
    if ficha.preco_venda > 0:
        resumo_data.append([f"PREÇO DE VENDA", f"R$ {float(ficha.preco_venda):,.2f}"])
    
    resumo_table = Table(resumo_data, colWidths=[10*cm, 6*cm])
    resumo_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#A5D6A7')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elementos.append(resumo_table)
    
    doc.build(elementos)
    
    if isinstance(output_stream, io.BytesIO):
        output_stream.seek(0)
        return output_stream.getvalue()
    
    return None
